from flask import Flask, render_template, request, redirect, url_for, jsonify, send_from_directory
import fitz # Libreria de Pypmupdf
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from dotenv import load_dotenv
from google import genai

load_dotenv()

# Configuración de la aplicación Flask
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

# Configuración del cliente de Gemini AI
cliente = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Lista para almacenar los datos extraídos de los PDFs
data_extraida = []

# Función para extraer metadatos y texto de un PDF, y generar un resumen con IA
def extraer_data_pdf(urlarchivo):
    """Extrae metadatos y texto del PDF, y genera un resumen con IA."""
    urlarchivo = os.path.abspath(urlarchivo)
    doc = fitz.open(urlarchivo)
    paginas_count  = doc.page_count
    imagenes_count = sum(len(page.get_images()) for page in doc)
    data_extraida_file = {}
    data_extraida_file["pais"] = ""
    data_extraida_file["paginas_imagenes"] = f"{paginas_count} / {imagenes_count}"
    metadata = doc.metadata
    data_extraida_file["nombre_archivo"] = os.path.basename(urlarchivo)
    data_extraida_file["titulo"] = metadata.get("title", "No detectado")
    data_extraida_file["autores"] = metadata.get("author", "No detectado")  
    data_extraida_file["palabras"] = metadata.get("keywords", "No detectado")
    data_extraida_file["tema"] = metadata.get("subject", "No detectado")
    data_extraida_file["paginas_imagenes"]  = f"{paginas_count} / {imagenes_count}"
    texto_completo = ""  # Variable para almacenar todo el texto del PDF
    for page_num in range(len(doc)):  # Itera sobre todas las páginas
        texto_completo += doc[page_num].get_text()
    
    # Limitar la longitud del texto para evitar errores de la API
    max_texto_longitud = 10000
    texto_resumir = texto_completo[:max_texto_longitud] if len(texto_completo) > max_texto_longitud else texto_completo

    try:
        # Generar el resumen usando la API de Gemini
        respuesta = cliente.models.generate_content(
            model="gemini-2.0-flash",
            contents=f"Resumen este texto en un solo parrafo en español: {texto_resumir}"
        )
        data_extraida_file["resumen"] = respuesta.text
        
        # Generar la detección del país usando la API de Gemini
        respuesta = cliente.models.generate_content(
            model="gemini-2.0-flash",
            contents=f"¿De qué país es este texto? Solo dime exactamente el país en ingles {texto_completo}"
        )
        data_extraida_file["pais"] = respuesta.text.strip() if respuesta.text else "No detectado"
        
    except Exception as e:
        data_extraida_file["resumen"] = f"Error al generar resumen: {e}"

    deteccion_anio = re.search(r'\b(20[0-3][0-9])\b', texto_completo) #se busca en el texto completo
    data_extraida_file["anio"] = deteccion_anio.group(1) if deteccion_anio else "No detectado"
    doc.close()
    return data_extraida_file

# Ruta para renderizar la página de estadísticas
@app.route("/estadisticas", methods=["GET", "POST"])
def estadisticas():
    if request.method == "GET":
        return render_template("estadisticas.html")
    return redirect(url_for("home"))


# Ruta para obtener los datos estadísticos extraídos de los PDFs
@app.route("/dataestadistica", methods=["GET"])
def data_estadistica():
    global data_extraida
    print("Datos extraídos:", data_extraida)
    return jsonify(
        data = data_extraida,
    )

# Ruta para cada vez que se presione el botón de "Buscar" se ejecutará la busqueda entre los PDFs local de la carpeta uploads
@app.route("/buscar", methods=["POST", "GET"])
def buscar_titulo():
    global data_extraida
    data_extraida = []  # Reiniciar la lista para la nueva búsqueda
    if request.method == "POST":
        termino_busqueda = request.form.get("buscar_tema", "").lower()
        if termino_busqueda:
            for nombrearchivo in os.listdir(app.config['UPLOAD_FOLDER']):
                if nombrearchivo.endswith(".pdf"):
                    urlarchivo = os.path.join(app.config['UPLOAD_FOLDER'], nombrearchivo)
                    data = extraer_data_pdf(urlarchivo)
                    if data and data["titulo"].lower().find(termino_busqueda) != -1:
                        data_extraida.append(data)
            termino_busqueda = None
            return render_template("index.html", data_list=data_extraida)
    return redirect(url_for("home", error="Por favor, ingresa un término de búsqueda."))

# Ruta para la página de inicio
@app.route("/", methods=["GET", "POST"])
def home():
    global data_extraida
    data_extraida = []
    return render_template('index.html')

# Ruta para la logica y exportar los datos a un archivo Excel
@app.route("/exportar_excel", methods=["POST"])
def exportar():
    global data_extraida
    if not data_extraida:
        return jsonify(error="No hay datos para exportar. Por favor, sube archivos PDF primero.")
    # Recibir las columnas seleccionadas desde el formulario
    columnas_seleccionadas = request.form.getlist('columnas')
    print("Columnas seleccionadas:", columnas_seleccionadas)
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos extraídos"

    # Definir un mapeo entre los nombres de las columnas y las claves del diccionario data
    mapeo_columnas = {
        "titulo": "titulo",
        "autores": "autores",
        "anio": "anio",
        "tema": "tema",
        "pais": "pais", 
        "palabras": "palabras",
        "paginas_imagenes":  "paginas_imagenes",
        "resumen": "resumen"
        
    }

    # Filtrar los encabezados basados en las columnas seleccionadas
    encabezado = [col for col in ["Titulo", "Autores", "Anio", "Tema","Pais", "Palabras","Paginas_Imagenes", "Resumen"] if mapeo_columnas[col.lower()] in columnas_seleccionadas]
    ws.append(encabezado)

    relleno_encabezado = PatternFill(start_color="002147", end_color="002147", fill_type="solid")
    fuente_encabezado = Font(color="FFFFFF", bold=True)
    borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col in range(1, len(encabezado) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = relleno_encabezado
        cell.border = borde
        cell.font = fuente_encabezado
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for data in data_extraida:
        # Filtrar los datos de cada fila según las columnas seleccionadas
        fila_data = [data[mapeo_columnas[col.lower()]] for col in encabezado]
        ws.append(fila_data)

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(encabezado)):
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(horizontal="left", vertical="top")

    nombrearchivo = "datos.xlsx"
    urlarchivo = os.path.join(app.config['UPLOAD_FOLDER'], nombrearchivo)
    wb.save(urlarchivo)
    file_url = url_for('descargar_excel')
    return jsonify(url=file_url)

# Ruta para descargar el archivo Excel generado
@app.route("/descargar_excel")
def descargar_excel():
    filename = "datos.xlsx"
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)