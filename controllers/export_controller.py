# Importación de librerías necesarias para la exportación y descarga de archivos
from flask import Blueprint, request, jsonify, url_for, send_from_directory, current_app
from openpyxl import Workbook                      # Para crear archivos Excel
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment  # Estilos para Excel
import os
from controllers import shared_data as datos_compartidos # Importa los datos compartidos de la aplicación

# Crear un blueprint específico para funcionalidades de exportación
export_bp = Blueprint('export', __name__)

# Ruta que genera un archivo Excel con los datos seleccionados por el usuario
@export_bp.route("/exportar_excel", methods=["POST"])
def exportar():
    # Obtiene los datos extraídos que se compartieron desde otros módulos
    datos_extraidos = datos_compartidos.get_datos_extraidos()
    print(datos_extraidos)
    # Validar que haya datos disponibles para exportar
    if not datos_extraidos:
        return jsonify(error="No hay datos para exportar.")

    # Obtener la lista de columnas seleccionadas por el usuario en el formulario
    columnas_seleccionadas = request.form.getlist('columnas')

    # Crear un nuevo libro de trabajo de Excel
    libro_trabajo = Workbook()
    hoja_activa = libro_trabajo.active
    hoja_activa.title = "Datos extraídos"

    # Mapeo de nombres de columnas del frontend a las claves del diccionario de datos
    mapeo_columnas = {
        "titulo": "titulo",
        "autores": "autores",
        "anio": "anio",
        "tema": "tema",
        "pais": "pais",
        "palabras": "palabras",
        "resumen": "resumen",
        "paginas_imagenes":"paginas_imagenes",
    }

    # Construir el encabezado del Excel con las columnas que el usuario seleccionó
    encabezado = [col.capitalize() for col in mapeo_columnas if col in columnas_seleccionadas]
    hoja_activa.append(encabezado)

    # Definir estilos para el encabezado de la hoja de Excel
    estilo_encabezado = {
        "relleno": PatternFill(start_color="002147", end_color="002147", fill_type="solid"),
        "fuente": Font(color="FFFFFF", bold=True),
        "borde": Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        "alineacion": Alignment(horizontal="center", vertical="center")
    }

    # Aplicar los estilos definidos a cada celda del encabezado
    for columna_num in range(1, len(encabezado) + 1):
        celda = hoja_activa.cell(row=1, column=columna_num)
        celda.fill = estilo_encabezado["relleno"]
        celda.border = estilo_encabezado["borde"]
        celda.font = estilo_encabezado["fuente"]
        celda.alignment = estilo_encabezado["alineacion"]

    # Llenar las filas de la hoja con los datos extraídos
    for dato in datos_extraidos:
        # Crea una fila solo con los datos de las columnas seleccionadas
        fila = [dato.get(mapeo_columnas[col.lower()], "") for col in encabezado]
        hoja_activa.append(fila)

    # Aplicar estilos a las celdas que contienen los datos
    for fila_celdas in hoja_activa.iter_rows(min_row=2):
        for celda in fila_celdas:
            celda.border = estilo_encabezado["borde"]
            celda.alignment = Alignment(horizontal="left", vertical="top")

    # Guardar el archivo Excel generado en la carpeta de subidas
    nombre_archivo = "datos.xlsx"
    ruta_archivo = os.path.join(current_app.config['UPLOAD_FOLDER'], nombre_archivo)
    libro_trabajo.save(ruta_archivo)

    # Devolver la URL para que el frontend pueda iniciar la descarga del archivo
    return jsonify(url=url_for("export.descargar_excel"))

# Ruta que permite al usuario descargar el archivo Excel que se generó previamente
@export_bp.route("/descargar_excel")
def descargar_excel():
    # Envía el archivo desde el directorio de subidas como un anexo para descargar
    return send_from_directory(
        current_app.config['UPLOAD_FOLDER'],
        "datos.xlsx",
        as_attachment=True  # Indica que el archivo debe descargarse en lugar de mostrarse en el navegador
    )
